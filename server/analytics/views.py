import io
from collections import defaultdict

from django.db.models import Count, Q, Sum, F, Subquery, OuterRef
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from bot2.models import Bot2Student, Bot2SurveyResponse, ProgramEnrollment, StudentRoster
from common.exceptions import build_error_response
from common.permissions import IsViewerOrAdminReadOnly
from common.time import parse_iso_datetime


BOT2_COURSE_YEARS = [1, 2, 3, 4, 5]


def _bot2_roster_totals_qs(campaign: str, course_year: int | None = None):
    qs = StudentRoster.objects.filter(is_active=True, roster_campaign=campaign)
    if course_year is not None:
        qs = qs.filter(course_year=course_year)
    return qs


def _resolve_academic_year(campaign: str, academic_year: str | None) -> str | None:
    """Return explicit academic_year or auto-detect latest from ProgramEnrollment."""
    if academic_year:
        return academic_year
    latest = (
        ProgramEnrollment.objects.filter(is_active=True, campaign=campaign)
        .order_by("-academic_year")
        .values_list("academic_year", flat=True)
        .first()
    )
    return latest


def _require_range(request):
    from_str = request.query_params.get("from")
    to_str = request.query_params.get("to")
    if not from_str or not to_str:
        return None, None, build_error_response("TIME_RANGE_REQUIRED", "from/to query params are required.", status.HTTP_400_BAD_REQUEST)
    start = parse_iso_datetime(from_str)
    end = parse_iso_datetime(to_str)
    if not start or not end:
        return None, None, build_error_response("INVALID_TIME_RANGE", "from/to must be ISO datetime.", status.HTTP_400_BAD_REQUEST)
    if start >= end:
        return None, None, build_error_response("INVALID_TIME_RANGE", "from must be earlier than to.", status.HTTP_400_BAD_REQUEST)
    return start, end, None


def _latest_responses_qs(start, end, campaign):
    """
    Return latest survey response per student within time range and campaign.
    Uses submitted_at for ordering, then created_at/id as tiebreakers.
    """
    base = Bot2SurveyResponse.objects.filter(
        submitted_at__gte=start,
        submitted_at__lte=end,
        survey_campaign=campaign,
    )
    latest_ids = (
        base.filter(student_id=OuterRef("student_id"))
        .order_by("-submitted_at", "-created_at", "-id")
        .values("id")[:1]
    )
    return base.annotate(latest_id=Subquery(latest_ids)).filter(id=F("latest_id"))


def _coverage_percent(responded, total):
    """Coverage %, clamped to [0, 100]. Totals and responses come from
    independent sources, so off-roster responders could otherwise push it >100."""
    if not total:
        return 0.0
    return round(min((responded or 0) * 100.0 / total, 100.0), 2)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def bot2_course_year_coverage(request):
    start, end, error = _require_range(request)
    if error:
        return error
    campaign = request.query_params.get("campaign", "default")
    academic_year = _resolve_academic_year(campaign, request.query_params.get("academic_year"))

    total_map = {}

    if academic_year:
        enroll_qs = ProgramEnrollment.objects.filter(is_active=True, campaign=campaign, academic_year=academic_year)
        totals = enroll_qs.values("course_year").annotate(total_students=Sum("student_count"))
        total_map.update({row["course_year"]: (row["total_students"] or 0) for row in totals})

        # ProgramEnrollment currently tracks 1-4. For graduates (5), fall back to roster counts.
        total_map[5] = _bot2_roster_totals_qs(campaign=campaign, course_year=5).count()
    else:
        totals = (
            _bot2_roster_totals_qs(campaign=campaign)
            .values("course_year")
            .annotate(total_students=Count("id"))
        )
        total_map.update({row["course_year"]: (row["total_students"] or 0) for row in totals})

    responded = (
        _latest_responses_qs(start, end, campaign)
        .values("course_year")
        .annotate(count=Count("student_id", distinct=True))
    )
    resp_map = {row["course_year"]: row["count"] for row in responded}

    result = []
    for year in BOT2_COURSE_YEARS:
        total = total_map.get(year, 0)
        resp = resp_map.get(year, 0)
        coverage = _coverage_percent(resp, total)
        result.append({"course_year": year, "total": total, "responded": resp, "coverage_percent": coverage})
    return Response(result)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def bot2_program_coverage(request):
    start, end, error = _require_range(request)
    if error:
        return error
    campaign = request.query_params.get("campaign", "default")
    academic_year = _resolve_academic_year(campaign, request.query_params.get("academic_year"))
    course_year = request.query_params.get("course_year")

    # Per-program totals. ProgramEnrollment tracks only years 1-4, so graduates
    # (year 5) always come from roster counts; accumulate per program.
    total_map = {}

    def _add_total(pid, name, value):
        if pid in total_map:
            total_map[pid]["total"] += value or 0
        else:
            total_map[pid] = {"program__name": name, "total": value or 0}

    if academic_year and course_year != "5":
        enroll_qs = ProgramEnrollment.objects.filter(is_active=True, campaign=campaign, academic_year=academic_year)
        if course_year:
            enroll_qs = enroll_qs.filter(course_year=course_year)
        for row in enroll_qs.values("program__id", "program__name").annotate(total=Sum("student_count")):
            _add_total(row["program__id"], row["program__name"], row["total"])
        # Fold in graduates (year 5) from the roster unless filtered to a 1-4 year.
        if not course_year:
            for row in (
                _bot2_roster_totals_qs(campaign=campaign, course_year=5)
                .values("program__id", "program__name")
                .annotate(total=Count("id"))
            ):
                _add_total(row["program__id"], row["program__name"], row["total"])
    else:
        roster_qs = _bot2_roster_totals_qs(campaign=campaign)
        if course_year:
            roster_qs = roster_qs.filter(course_year=course_year)
        for row in roster_qs.values("program__id", "program__name").annotate(total=Count("id")):
            _add_total(row["program__id"], row["program__name"], row["total"])

    resp_qs = _latest_responses_qs(start, end, campaign)
    if course_year:
        resp_qs = resp_qs.filter(course_year=course_year)
    resp_map = {
        r["program__id"]: r
        for r in resp_qs.values("program__id", "program__name").annotate(count=Count("student_id", distinct=True))
    }

    data = []
    # Iterate the union of totals and responses so programs that have responses
    # but no enrollment row (graduates / off-roster) are not silently dropped.
    for program_id in total_map.keys() | resp_map.keys():
        total = total_map.get(program_id, {}).get("total", 0)
        name = (total_map.get(program_id) or resp_map.get(program_id, {})).get("program__name")
        resp_row = resp_map.get(program_id)
        resp = resp_row["count"] if resp_row else 0
        data.append(
            {
                "program_id": program_id,
                "program_name": name,
                "total": total,
                "responded": resp,
                "coverage_percent": _coverage_percent(resp, total),
            }
        )
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def bot2_program_course_matrix(request):
    start, end, error = _require_range(request)
    if error:
        return error
    campaign = request.query_params.get("campaign", "default")
    academic_year = _resolve_academic_year(campaign, request.query_params.get("academic_year"))

    if academic_year:
        enroll_qs = ProgramEnrollment.objects.filter(is_active=True, campaign=campaign, academic_year=academic_year)
        totals = enroll_qs.values("program__id", "program__name", "course_year").annotate(total=Sum("student_count"))

        # ProgramEnrollment currently tracks 1-4. For graduates (5), fall back to roster counts.
        roster_totals = (
            _bot2_roster_totals_qs(campaign=campaign, course_year=5)
            .values("program__id", "program__name", "course_year")
            .annotate(total=Count("id"))
        )
    else:
        totals = (
            _bot2_roster_totals_qs(campaign=campaign)
            .values("program__id", "program__name", "course_year")
            .annotate(total=Count("id"))
        )
        roster_totals = []
    responded = (
        _latest_responses_qs(start, end, campaign)
        .values("program__id", "program__name", "course_year")
        .annotate(count=Count("student_id", distinct=True))
    )

    programs = {}
    totals_map = defaultdict(dict)
    for row in totals:
        programs[row["program__id"]] = row["program__name"]
        totals_map[row["program__id"]][row["course_year"]] = row["total"]

    # Add graduates (course_year=5) totals from roster
    for row in roster_totals:
        programs[row["program__id"]] = row["program__name"]
        totals_map[row["program__id"]][row["course_year"]] = row["total"]

    resp_map = defaultdict(dict)
    for row in responded:
        programs[row["program__id"]] = row["program__name"]
        resp_map[row["program__id"]][row["course_year"]] = row["count"]

    program_list = [{"id": pid, "name": name} for pid, name in programs.items()]
    cells = []
    for pid, name in programs.items():
        for year in BOT2_COURSE_YEARS:
            total = totals_map[pid].get(year, 0)
            resp = resp_map[pid].get(year, 0)
            coverage = _coverage_percent(resp, total)
            cells.append(
                {
                    "program_id": pid,
                    "course_year": year,
                    "total": total,
                    "responded": resp,
                    "coverage_percent": coverage,
                }
            )

    return Response({"years": BOT2_COURSE_YEARS, "programs": program_list, "cells": cells})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def bot2_program_details_by_year(request):
    """Get program breakdown for a specific course year with employment stats."""
    start, end, error = _require_range(request)
    if error:
        return error
    campaign = request.query_params.get("campaign", "default")
    academic_year = _resolve_academic_year(campaign, request.query_params.get("academic_year"))
    course_year = request.query_params.get("course_year")
    
    if not course_year:
        return build_error_response("COURSE_YEAR_REQUIRED", "course_year query param is required.", status.HTTP_400_BAD_REQUEST)
    
    try:
        course_year = int(course_year)
    except ValueError:
        return build_error_response("INVALID_COURSE_YEAR", "course_year must be an integer.", status.HTTP_400_BAD_REQUEST)

    if academic_year and course_year != 5:
        enroll_qs = ProgramEnrollment.objects.filter(
            is_active=True, campaign=campaign, academic_year=academic_year, course_year=course_year
        )
        totals = enroll_qs.values("program__id", "program__name").annotate(total=Sum("student_count"))
    else:
        totals = (
            _bot2_roster_totals_qs(campaign=campaign, course_year=course_year)
            .values("program__id", "program__name")
            .annotate(total=Count("id"))
        )
    total_map = {row["program__id"]: {"name": row["program__name"], "total": row["total"] or 0} for row in totals}
    
    # Get survey responses per program for this year (count unique students)
    responded = (
        _latest_responses_qs(start, end, campaign)
        .filter(course_year=course_year)
        .values("program__id", "program__name")
        .annotate(count=Count("student_id", distinct=True))
    )
    
    # Employment status breakdown (unique students)
    employment = (
        _latest_responses_qs(start, end, campaign)
        .filter(course_year=course_year)
        .values("program__id", "employment_status")
        .annotate(count=Count("student_id", distinct=True))
    )
    
    # Add responded counts
    for row in responded:
        if row["program__id"] in total_map:
            total_map[row["program__id"]]["responded"] = row["count"]
        else:
            total_map[row["program__id"]] = {"name": row["program__name"], "total": 0, "responded": row["count"]}
    
    # Add employment breakdown
    for program_id in total_map:
        total_map[program_id]["employed"] = 0
        total_map[program_id]["unemployed"] = 0
    
    for row in employment:
        if row["program__id"] in total_map:
            emp_status = row["employment_status"].lower() if row["employment_status"] else ""
            if "ishlayapman" in emp_status or "employed" in emp_status or "ишлаяпман" in emp_status:
                total_map[row["program__id"]]["employed"] += row["count"]
            else:
                total_map[row["program__id"]]["unemployed"] += row["count"]

    # Format response
    data = []
    for program_id, info in total_map.items():
        total = info.get("total", 0)
        responded = info.get("responded", 0)
        coverage = _coverage_percent(responded, total)
        data.append({
            "program_id": program_id,
            "program_name": info["name"],
            "total": total,
            "responded": responded,
            "coverage_percent": coverage,
            "employed": info.get("employed", 0),
            "unemployed": info.get("unemployed", 0),
        })
    
    # Sort by total students descending
    data.sort(key=lambda x: x["total"], reverse=True)
    
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def enrollments_overview(request):
    """
    Aggregate ProgramEnrollment by program and course year with coverage metrics.
    Requires time range to align with survey responses.
    """
    start, end, error = _require_range(request)
    if error:
        return error

    campaign = request.query_params.get("campaign", "default")
    academic_year = _resolve_academic_year(campaign, request.query_params.get("academic_year"))

    if academic_year:
        enroll_qs = ProgramEnrollment.objects.filter(is_active=True, campaign=campaign, academic_year=academic_year)
        totals = list(
            enroll_qs.values("program__id", "program__name", "course_year")
            .annotate(total=Sum("student_count"))
            .order_by("program__name", "course_year")
        )

        # Add graduates (course_year=5) totals from roster counts.
        totals.extend(
            list(
                _bot2_roster_totals_qs(campaign=campaign, course_year=5)
                .values("program__id", "program__name", "course_year")
                .annotate(total=Count("id"))
                .order_by("program__name", "course_year")
            )
        )
    else:
        totals = list(
            _bot2_roster_totals_qs(campaign=campaign)
            .values("program__id", "program__name", "course_year")
            .annotate(total=Count("id"))
            .order_by("program__name", "course_year")
        )

    responded = (
        _latest_responses_qs(start, end, campaign)
        .values("program__id", "program__name", "course_year")
        .annotate(count=Count("student_id", distinct=True))
    )

    resp_map = {}
    for row in responded:
        key = (row["program__id"], row["course_year"])
        resp_map[key] = row["count"]

    overview = []
    total_students = 0
    total_responded = 0
    yearly = {year: {"total": 0, "responded": 0} for year in BOT2_COURSE_YEARS}

    for row in totals:
        program_id = row["program__id"]
        course_year = row["course_year"]
        total = row["total"] or 0
        responded_count = resp_map.get((program_id, course_year), 0)
        coverage = _coverage_percent(responded_count, total)

        overview.append(
            {
                "program_id": program_id,
                "program_name": row["program__name"],
                "course_year": course_year,
                "total": total,
                "responded": responded_count,
                "coverage_percent": coverage,
            }
        )

        total_students += total
        total_responded += responded_count
        if course_year in yearly:
            yearly[course_year]["total"] += total
            yearly[course_year]["responded"] += responded_count

    yearly_list = []
    for year, vals in yearly.items():
        total = vals["total"]
        resp = vals["responded"]
        yearly_list.append(
            {
                "course_year": year,
                "total": total,
                "responded": resp,
                "coverage_percent": _coverage_percent(resp, total),
            }
        )

    overall_coverage = _coverage_percent(total_responded, total_students)

    return Response(
        {
            "total_students": total_students,
            "total_responded": total_responded,
            "coverage_percent": overall_coverage,
            "by_year": yearly_list,
            "by_program": overview,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def bot2_academic_years(request):
    """List distinct academic_year values from ProgramEnrollment, newest first."""
    campaign = request.query_params.get("campaign", "default")
    years = (
        ProgramEnrollment.objects.filter(is_active=True, campaign=campaign)
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )
    return Response(list(years))


def _students_by_direction_data(campaign: str) -> list:
    """
    Faza F: Students grouped by direction/program.
    - total: StudentRoster count per program
    - registered: Bot2Student count (has telegram_user_id)
    - employed: latest survey per student where employment_status looks like "employed"
    """
    # Roster totals per program
    totals = (
        StudentRoster.objects.filter(is_active=True, roster_campaign=campaign)
        .values("program_id", "program__name", "program__name_uz", "program__name_ru")
        .annotate(total=Count("id"))
    )
    total_map = {
        row["program_id"]: {
            "program_name": row["program__name"],
            "program_name_uz": row["program__name_uz"],
            "program_name_ru": row["program__name_ru"],
            "total": row["total"],
        }
        for row in totals
    }

    # Registered students per program (has telegram_user_id — interacted with bot)
    registered = (
        Bot2Student.objects.filter(
            roster__is_active=True,
            roster__roster_campaign=campaign,
            telegram_user_id__isnull=False,
        )
        .values("roster__program_id")
        .annotate(count=Count("id"))
    )
    registered_map = {row["roster__program_id"]: row["count"] for row in registered}

    # Employed: latest survey per student, check employment_status
    latest_ids = (
        Bot2SurveyResponse.objects
        .filter(student_id=OuterRef("student_id"), submitted_at__isnull=False)
        .order_by("-submitted_at", "-created_at")
        .values("id")[:1]
    )
    employed_count = (
        Bot2SurveyResponse.objects
        .annotate(latest_id=Subquery(latest_ids))
        .filter(id=F("latest_id"))
        .filter(
            Q(employment_status__icontains="ishlayapman")
            | Q(employment_status__icontains="employed")
            | Q(employment_status__icontains="ишлаяпман")
        )
        .values("program_id")
        .annotate(count=Count("student_id", distinct=True))
    )
    employed_map = {row["program_id"]: row["count"] for row in employed_count}

    result = []
    for program_id, info in total_map.items():
        result.append({
            "program_id": program_id,
            "program_name": info["program_name"],
            "program_name_uz": info["program_name_uz"],
            "program_name_ru": info["program_name_ru"],
            "total": info["total"],
            "registered": registered_map.get(program_id, 0),
            "employed": employed_map.get(program_id, 0),
        })

    result.sort(key=lambda x: x["total"], reverse=True)
    return result


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def students_by_direction(request):
    """GET /api/v1/analytics/students-by-direction — per-program totals."""
    campaign = request.query_params.get("campaign", "default")
    return Response(_students_by_direction_data(campaign))


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsViewerOrAdminReadOnly])
def students_by_direction_xlsx(request):
    """GET /api/v1/analytics/students-by-direction.xlsx — openpyxl export."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    campaign = request.query_params.get("campaign", "default")
    rows = _students_by_direction_data(campaign)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students by Direction"

    headers = ["Program", "Total", "Registered", "Employed", "Registered %", "Employed %"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        total = row["total"] or 0
        registered = row["registered"] or 0
        employed = row["employed"] or 0
        reg_pct = round(registered * 100.0 / total, 1) if total else 0.0
        emp_pct = round(employed * 100.0 / total, 1) if total else 0.0

        ws.cell(row=row_idx, column=1, value=row["program_name"])
        ws.cell(row=row_idx, column=2, value=total)
        ws.cell(row=row_idx, column=3, value=registered)
        ws.cell(row=row_idx, column=4, value=employed)
        ws.cell(row=row_idx, column=5, value=reg_pct)
        ws.cell(row=row_idx, column=6, value=emp_pct)

    ws.column_dimensions["A"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="students-by-direction-{campaign}.xlsx"'
    return response
